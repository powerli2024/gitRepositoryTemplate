#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""v1 多标签关键词 vs v2 互斥主类：完善 linguistic_pattern_analysis.xlsx。"""
from __future__ import annotations

import json
import random
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PUNCT = str.maketrans("", "", "，。！？、；：\"\"''「」『』（）【】《》…·—–-.,!?;:\"'`")
DIGIT = re.compile(r"\d|[零一二两三四五六七八九十百千万半]")

# 句式动词（不写设备名，避免记本库词表）
ACT_VERBS = [
    "打开", "开启", "关闭", "关掉", "关上", "开机", "关机",
    "播放", "暂停", "停止", "启动", "开始", "切到", "切换",
    "拉起", "拉开", "拉上", "放下", "降下", "升起", "恢复", "设置",
    "帮我", "请帮我", "麻烦", "取消", "给我", "遮上",
    "开开", "开下", "关下",
]
# 单字动词只在句首计（开冰箱门 / 关下空调），避免「关于」
ACT_PREFIX = ("开", "关", "放", "换")
STATE_MARK = [
    "调到", "调成", "调为", "调高", "调低", "调亮", "调暗", "调大", "调小",
    "小一点", "大一点", "亮一点", "暗一点", "高一点", "低一点",
    "最大", "最小", "最亮", "最暗", "最低", "最高",
    "加大", "减小", "全开", "半开", "停下",
]
Q_WORDS = ["什么", "怎么", "怎样", "为何", "为什么", "哪", "几", "是否", "吗"]
LIFE_MARK = re.compile(
    r"我(要|想|准备|马上)|回家(了|啦|咯)|出门(了|啦|咯)|我回来|"
    r"有点(热|冷|饿)|吃(完|饱|饭)了"
)
TTS_MARK = ["语音功能已关闭", "主人当前", "可以使用遥控器"]
DISCOURSE = ["就是", "因为", "我觉得", "那我们", "怪不得", "是吧", "待会", "这个那个"]
PARTICLES = ["吧", "啊", "呃", "嗯", "哈", "嘛", "呀", "哦", "喔", "咯", "啦"]
COMPOUND_LINK = ["然后", "还有", "接着", "再"]

V1_LABELS = ["动作命令", "状态调节", "模式/场景", "信息询问", "生活陈述/社交", "复合命令", "闲聊/背景"]


def nfkc(s: Any) -> str:
    if not s:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    t = "".join(ch for ch in t if not ch.isspace()).translate(PUNCT)
    return t.lower().strip()


def count_verbs(t: str) -> int:
    n, i = 0, 0
    verbs = sorted(ACT_VERBS, key=len, reverse=True)
    while i < len(t):
        hit = next((v for v in verbs if t.startswith(v, i)), None)
        if hit:
            n += 1
            i += len(hit)
        elif i == 0 and t.startswith(ACT_PREFIX) and len(t) >= 3 and not t.startswith("关于"):
            n += 1
            i += 1
        else:
            i += 1
    return n


def v1_tags(t: str) -> set[str]:
    """复现原表：多标签关键词，语气词可与命令共存。"""
    tags: set[str] = set()
    if any(v in t for v in ACT_VERBS) or t.startswith("播放"):
        tags.add("动作命令")
    if any(s in t for s in STATE_MARK) or (DIGIT.search(t) and any(x in t for x in ("度", "百分", "风速", "亮度", "温度", "音量"))):
        tags.add("状态调节")
    if "模式" in t or "场景" in t:
        tags.add("模式/场景")
    if any(q in t for q in Q_WORDS):
        tags.add("信息询问")
    if LIFE_MARK.search(t):
        tags.add("生活陈述/社交")
    if count_verbs(t) >= 2 or any(x in t for x in COMPOUND_LINK) and count_verbs(t) >= 1:
        tags.add("复合命令")
    if any(x in t for x in DISCOURSE) or any(p in t for p in PARTICLES):
        tags.add("闲聊/背景")
    return tags


def v1_binary_command(tags: set[str]) -> bool:
    core = tags & {"动作命令", "状态调节", "模式/场景", "复合命令"}
    return bool(core) and "闲聊/背景" not in tags


def classify_v2(raw: str) -> dict[str, Any]:
    t = nfkc(raw)
    n_verb = count_verbs(t)
    has_state = any(s in t for s in STATE_MARK) or (
        bool(DIGIT.search(t)) and any(x in t for x in ("度", "百分", "风速", "亮度", "温度", "音量", "色温"))
    )
    has_mode = "模式" in t or "场景" in t
    has_q = any(q in t for q in ("什么", "怎么", "怎样", "为何", "为什么", "哪", "是否"))
    has_ma = t.endswith("吗")
    has_life = bool(LIFE_MARK.search(t))
    has_tts = any(x in t for x in TTS_MARK)
    has_disc = any(x in t for x in DISCOURSE)
    has_part = any(p in t for p in PARTICLES)
    has_link = any(x in t for x in COMPOUND_LINK)
    command_frame = n_verb >= 1 or has_state or has_mode

    tags: list[str] = []
    if has_tts:
        tags.append("设备播报")
    if has_q or (has_ma and not command_frame):
        tags.append("信息询问")
    if n_verb >= 2 or (has_link and (n_verb + int(has_state) + int(has_mode)) >= 2):
        tags.append("复合命令")
    if has_state:
        tags.append("状态调节")
    if n_verb >= 1:
        tags.append("动作命令")
    if has_mode:
        tags.append("模式/场景")
    if has_life:
        tags.append("生活陈述/社交")
    if has_part:
        tags.append("带语气词")
    if has_disc:
        tags.append("带话语标记")

    # 互斥主类：问句 > 播报 > 复合 > 状态 > 动作 > 光杆模式 > 生活陈述 > 闲聊 > 其他
    if has_tts:
        primary = "设备播报"
    elif has_q:
        primary = "信息询问"
    elif n_verb >= 2 or (has_link and (n_verb + int(has_state) + int(has_mode)) >= 2):
        primary = "复合命令"
    elif has_state:
        primary = "状态调节"
    elif n_verb >= 1:
        primary = "动作命令"
    elif has_mode:
        primary = "模式/场景"
    elif has_life:
        primary = "生活陈述/社交"
    elif has_disc or (len(t) >= 16 and not command_frame) or (len(t) <= 3 and not command_frame):
        primary = "闲聊/背景"
    elif not t:
        primary = "空转写"
    elif 4 <= len(t) <= 10 and not has_disc:
        # 省略指令：防直吹、风速最低、浪漫就餐 — 无完整动词但仍是面向设备的短句
        primary = "省略指令"
        tags.append("省略指令")
    else:
        primary = "其他"

    task = primary in {
        "动作命令", "状态调节", "模式/场景", "复合命令",
        "信息询问", "生活陈述/社交", "省略指令",
    }
    return {
        "primary": primary,
        "tags": tags,
        "n_verb": n_verb,
        "len": len(t),
        "has_particle": has_part,
        "command_frame": command_frame,
        "task_oriented": task,
        "text": t,
    }


def mean(xs: list[float]) -> float:
    return sum(xs) / len(xs) if xs else 0.0


def contest_of(rows, pred) -> dict[str, float]:
    pos_c, n_neg, n_rej, n_pos, n_fr = [], 0, 0, 0, 0
    for r in rows:
        rej = pred(r)
        if r["split"] == "pos":
            n_pos += 1
            if rej:
                n_fr += 1
                pos_c.append(1.0)
            else:
                pos_c.append(float(r["cer"] if r.get("cer") is not None else 1))
        else:
            n_neg += 1
            if rej:
                n_rej += 1
    rr = n_rej / n_neg if n_neg else 0
    frr = n_fr / n_pos if n_pos else 1
    cer = mean(pos_c)
    return {"rr": rr, "frr": frr, "cer": cer, "contest": 0.5 * rr + 0.5 * (1 - cer)}


def binary_metrics(y_true: list[int], y_pred: list[int]) -> dict[str, Any]:
    tp = sum(t and p for t, p in zip(y_true, y_pred))
    fp = sum((not t) and p for t, p in zip(y_true, y_pred))
    fn = sum(t and (not p) for t, p in zip(y_true, y_pred))
    tn = sum((not t) and (not p) for t, p in zip(y_true, y_pred))
    prec = tp / (tp + fp) if tp + fp else 0
    rec = tp / (tp + fn) if tp + fn else 0
    acc = (tp + tn) / len(y_true) if y_true else 0
    return {
        "TP": tp, "FP": fp, "FN": fn, "TN": tn,
        "精确率": round(prec, 4), "召回率": round(rec, 4), "准确率": round(acc, 4),
        "F1": round(2 * prec * rec / (prec + rec), 4) if prec + rec else 0,
    }


def dist(counter: Counter, n: int) -> list[list[Any]]:
    labels = [
        "动作命令", "状态调节", "模式/场景", "信息询问",
        "生活陈述/社交", "复合命令", "省略指令", "闲聊/背景", "设备播报", "其他", "空转写",
    ]
    rows = []
    for k in labels:
        c = counter.get(k, 0)
        if c or k in V1_LABELS or k in ("设备播报", "其他", "空转写"):
            rows.append([k, c, f"{c / n:.2%}" if n else "0%"])
    return rows


def main() -> int:
    root = Path(r"d:\media\datasetA\sssss")
    asr = json.loads((root / "no_sep.json").read_text(encoding="utf-8"))
    scores = {}
    for line in (root / "scores (1).jsonl").read_text(encoding="utf-8").splitlines():
        if line.strip():
            r = json.loads(line)
            scores[r["uid"]] = r

    for r in asr:
        cmd = nfkc(r.get("cmd_text"))
        hyp = nfkc(r.get("hyp_norm") or r.get("asr_text"))
        r["cmd_n"] = cmd
        r["hyp_n"] = hyp
        r["v1_cmd"] = v1_tags(cmd) if cmd else set()
        r["v1_hyp"] = v1_tags(hyp)
        r["v2_cmd"] = classify_v2(cmd) if r["split"] == "pos" else None
        r["v2_hyp"] = classify_v2(hyp)
        r["s"] = float(scores[r["uid"]]["presence_score"])
        r["lang"] = r.get("lang") or "zh"

    pos = [r for r in asr if r["split"] == "pos"]
    neg = [r for r in asr if r["split"] == "neg"]

    def multi_dist(key: str, rows: list) -> dict[str, int]:
        c: Counter[str] = Counter()
        for r in rows:
            for t in r[key]:
                c[t] += 1
        return dict(c)

    v1_cmd = multi_dist("v1_cmd", pos)
    v1_pos_hyp = multi_dist("v1_hyp", pos)
    v1_neg_hyp = multi_dist("v1_hyp", neg)

    v2_cmd = Counter(r["v2_cmd"]["primary"] for r in pos)
    v2_pos_hyp = Counter(r["v2_hyp"]["primary"] for r in pos)
    v2_neg_hyp = Counter(r["v2_hyp"]["primary"] for r in neg)

    overlap_v1_cmd = sum(1 for r in pos if len(r["v1_cmd"]) > 1) / len(pos)
    unlabeled_v1_hyp_pos = sum(1 for r in pos if not r["v1_hyp"]) / len(pos)
    unlabeled_v1_hyp_neg = sum(1 for r in neg if not r["v1_hyp"]) / len(neg)
    particle_override = sum(
        1 for r in pos
        if r["v2_hyp"]["command_frame"] and r["v2_hyp"]["has_particle"]
        and "闲聊/背景" in r["v1_hyp"]
    )

    # cmd vs hyp 主类一致率
    agree = sum(1 for r in pos if r["v2_cmd"]["primary"] == r["v2_hyp"]["primary"]) / len(pos)

    y_pos_hyp = [1] * len(pos) + [0] * len(neg)
    v1_pred = [int(v1_binary_command(r["v1_hyp"])) for r in pos + neg]
    v2_pred = [int(r["v2_hyp"]["task_oriented"]) for r in pos + neg]
    v1_m = binary_metrics(y_pos_hyp, v1_pred)
    v2_m = binary_metrics(y_pos_hyp, v2_pred)

    # 理想：pos cmd vs neg hyp
    v1_ideal = binary_metrics(
        [1] * len(pos) + [0] * len(neg),
        [int(v1_binary_command(r["v1_cmd"])) for r in pos] + [int(v1_binary_command(r["v1_hyp"])) for r in neg],
    )
    v2_ideal = binary_metrics(
        [1] * len(pos) + [0] * len(neg),
        [int(r["v2_cmd"]["task_oriented"]) for r in pos] + [int(r["v2_hyp"]["task_oriented"]) for r in neg],
    )

    # holdout overlay
    rng = random.Random(7)
    by: dict[tuple, list] = defaultdict(list)
    for r in asr:
        by[(r["split"], r["lang"])].append(r)
    train, test = [], []
    for g in by.values():
        g = list(g)
        rng.shuffle(g)
        n = max(1, int(round(len(g) * 0.3)))
        test.extend(g[:n])
        train.extend(g[n:])

    def lang_thr(subset):
        out = {}
        for lang in ("zh", "en"):
            ps = [r["s"] for r in subset if r["split"] == "pos" and r["lang"] == lang]
            ns = [r["s"] for r in subset if r["split"] == "neg" and r["lang"] == lang]
            best = None
            for thr in sorted(set(round(x, 4) for x in ps + ns)):
                frr = sum(x < thr for x in ps) / len(ps)
                rr = sum(x < thr for x in ns) / len(ns)
                pxy = 0.5 * rr + 0.5 * (1 - frr)
                if best is None or pxy > best[0]:
                    best = (pxy, thr)
            out[lang] = best[1]
        return out["zh"], out["en"]

    zt, et = lang_thr(train)

    def thr_of(r):
        return zt if r["lang"] == "zh" else et

    def score_only(r):
        return r["s"] < thr_of(r)

    def gray(r, g=0.05):
        return abs(r["s"] - thr_of(r)) <= g

    def overlay_len(r, L=16):
        return True if gray(r) and r["v2_hyp"]["len"] >= L else score_only(r)

    def overlay_v1(r):
        # 灰区：v1 判定非命令则拒
        return True if gray(r) and not v1_binary_command(r["v1_hyp"]) else score_only(r)

    def overlay_v2_nontask(r):
        return True if gray(r) and not r["v2_hyp"]["task_oriented"] else score_only(r)

    def overlay_v2_chat(r):
        return True if gray(r) and r["v2_hyp"]["primary"] in ("闲聊/背景", "设备播报", "其他", "空转写") else score_only(r)

    def overlay_len_or_nontask(r):
        h = r["v2_hyp"]
        hit = gray(r) and (h["len"] >= 16 or not h["task_oriented"])
        return True if hit else score_only(r)

    def overlay_len_and_nontask(r):
        h = r["v2_hyp"]
        hit = gray(r) and h["len"] >= 16 and not h["task_oriented"]
        return True if hit else score_only(r)

    hold_rows = []
    for name, fn in [
        ("纯余弦", score_only),
        ("灰区 + v1非命令", overlay_v1),
        ("灰区 + v2非任务", overlay_v2_nontask),
        ("灰区 + v2闲聊/播报/其他", overlay_v2_chat),
        ("灰区 + 长度≥16", overlay_len),
        ("灰区 + 长度或非任务", overlay_len_or_nontask),
        ("灰区 + 长度且非任务", overlay_len_and_nontask),
    ]:
        tr, te = contest_of(train, fn), contest_of(test, fn)
        hold_rows.append({
            "name": name,
            "train_C": round(tr["contest"], 4),
            "test_C": round(te["contest"], 4),
            "test_RR": round(te["rr"], 4),
            "test_FRR": round(te["frr"], 4),
            "test_CER": round(te["cer"], 4),
        })

    # 示例：v1 错成闲聊、v2 纠正
    fixed_particle = [
        {
            "uid": r["uid"],
            "cmd": r.get("cmd_text"),
            "asr": r.get("asr_text"),
            "v1": ",".join(sorted(r["v1_hyp"])),
            "v2": r["v2_hyp"]["primary"],
        }
        for r in pos
        if r["v2_hyp"]["command_frame"] and r["v2_hyp"]["has_particle"] and "闲聊/背景" in r["v1_hyp"]
    ][:8]

    still_hard_neg = [
        {"uid": r["uid"], "asr": r.get("asr_text"), "v2": r["v2_hyp"]["primary"], "len": r["v2_hyp"]["len"]}
        for r in neg if r["v2_hyp"]["task_oriented"] and r["v2_hyp"]["len"] < 16
    ][:8]

    compound_gain = [
        {"uid": r["uid"], "cmd": r.get("cmd_text"), "v1": ",".join(sorted(r["v1_cmd"])), "v2": r["v2_cmd"]["primary"]}
        for r in pos if r["v2_cmd"]["primary"] == "复合命令"
    ][:8]

    summary = {
        "n_pos": len(pos),
        "n_neg": len(neg),
        "v1_overlap_pos_cmd": round(overlap_v1_cmd, 4),
        "v1_unlabeled_pos_hyp": round(unlabeled_v1_hyp_pos, 4),
        "v1_unlabeled_neg_hyp": round(unlabeled_v1_hyp_neg, 4),
        "v1_particle_false_chat": particle_override,
        "v2_cmd_hyp_agree": round(agree, 4),
        "v1_asr_metrics": v1_m,
        "v2_asr_metrics": v2_m,
        "v1_ideal_metrics": v1_ideal,
        "v2_ideal_metrics": v2_ideal,
        "v1_cmd_multi": v1_cmd,
        "v2_cmd_primary": dict(v2_cmd),
        "v2_pos_hyp_primary": dict(v2_pos_hyp),
        "v2_neg_hyp_primary": dict(v2_neg_hyp),
        "holdout": hold_rows,
        "thrs": {"zh": zt, "en": et},
    }
    (root / "pattern_v2_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    # ---- xlsx ----
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)

    def autosize(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap

    # 0 修订说明
    ws = wb.active
    ws.title = "v1问题与v2修订"
    ws.append(["条目", "说明"])
    notes = [
        ["v1 多标签", "同一句可同时是动作命令+模式+闲聊。分布表各行之和不是 100%（pos cmd 合计约 107%）。"],
        ["语气词误伤", f"pos 中 {particle_override} 条有命令框架却因吧/啊/嗯被打上闲聊；v1 二分类「核心命令且非闲聊」会把它们丢掉。"],
        ["复合命令过窄", f"v1 复合仅 {v1_cmd.get('复合命令', 0)} 条；「打开烟机打开洗衣模式打开新风」被模式关键词抢走。v2 按动词次数≥2 提升为复合。"],
        ["动作 vs 生活陈述", "「我要出门了」不应算动作命令。v2：有打开/关/播放才是动作；我要/我想无设备动词才是生活陈述。"],
        ["打开X模式", "主类=动作命令，标记里保留模式/场景。光杆「日光模式」才是模式主类。"],
        ["信息询问", "仅强疑问词（什么/怎么/哪）。句尾「吗」不覆盖已有命令框架，避免「我要看电影吗」三标签。"],
        ["新增", "设备播报、省略指令、其他、空转写。语气词改为特征不是类。主类互斥合计 100%。"],
        ["拒识用法", "不要 7 类硬分类。灰区只用：非任务主类，或长度≥16。词表不当拒绝依据。"],
    ]
    for row in notes:
        ws.append(row)
    style_header(ws)
    autosize(ws, [18, 88])

    # 1 定义
    ws = wb.create_sheet("v2范式定义")
    ws.append(["主类（互斥）", "判定优先级与条件", "可叠加标记", "示例"])
    defs = [
        ["信息询问", "1. 含什么/怎么/怎样/哪/是否（不问句尾单独的吗）", "可叠加动作", "元宵节灯会什么时候；羊毛衫怎么洗"],
        ["设备播报", "2. 语音功能已关闭 / 主人当前 / 用遥控器", "—", "主人当前语音功能已关闭…"],
        ["复合命令", "3. 句式动词≥2，或「然后/还有」连接两个动作/调节", "动作+状态+模式", "打开烟机打开洗衣模式；调亮然后开窗帘"],
        ["状态调节", "4. 调到/调成/小一点/度/百分/风速…", "可叠加动作", "风量调到百分之三十"],
        ["动作命令", "5. 打开/关/播放/暂停/拉起/帮我…（含打开X模式）", "模式标记", "打开空调；打开回家模式吧"],
        ["模式/场景", "6. 光杆「X模式/X场景」，前面没有动作动词", "—", "日光模式；观影模式"],
        ["生活陈述/社交", "7. 我要/我想/准备/出门了/回家了，且没有设备动词", "—", "我要出门了；我准备做饭了"],
        ["省略指令", "8. 无完整动词、4–10 字、无话语标记：省略祈使/功能名", "—", "风速最低；防直吹；开冰箱门；加大风量"],
        ["闲聊/背景", "9. 就是/因为/那我们…，或无命令框架的过长/过短残片", "—", "就是他们的灯光会变化；这个灯"],
        ["其他 / 空转写", "10. 剩余", "—", "未覆盖的长陈述"],
    ]
    for row in defs:
        ws.append(row)
    style_header(ws)
    autosize(ws, [18, 55, 22, 42])

    # 2 分布对比
    ws = wb.create_sheet("主类分布对比")
    ws.append([
        "主类",
        "v2 pos cmd", "占比",
        "v2 pos asr", "占比",
        "v2 neg asr", "占比",
        "v1 pos cmd(多标签)", "v1 pos asr(多标签)", "v1 neg asr(多标签)",
    ])
    labels = [
        "动作命令", "状态调节", "模式/场景", "信息询问", "生活陈述/社交",
        "复合命令", "省略指令", "闲聊/背景", "设备播报", "其他", "空转写",
    ]
    for k in labels:
        pc, pa, na = v2_cmd.get(k, 0), v2_pos_hyp.get(k, 0), v2_neg_hyp.get(k, 0)
        ws.append([
            k, pc, f"{pc/len(pos):.2%}", pa, f"{pa/len(pos):.2%}", na, f"{na/len(neg):.2%}",
            v1_cmd.get(k, 0), v1_pos_hyp.get(k, 0), v1_neg_hyp.get(k, 0),
        ])
    ws.append([])
    ws.append(["诊断", "数值"])
    ws.append(["v1 pos cmd 多标签占比", f"{overlap_v1_cmd:.2%}"])
    ws.append(["v1 pos asr 未打上任何标签", f"{unlabeled_v1_hyp_pos:.2%}"])
    ws.append(["v1 neg asr 未打上任何标签", f"{unlabeled_v1_hyp_neg:.2%}"])
    ws.append(["命令框架+语气词却被 v1 标闲聊（pos asr）", particle_override])
    ws.append(["v2 pos 主类 cmd↔asr 一致率", f"{agree:.2%}"])
    style_header(ws)
    autosize(ws, [20, 14, 12, 14, 12, 14, 12, 20, 20, 20])

    # 3 判别效果
    ws = wb.create_sheet("任务导向判别")
    ws.append(["方案", "TP", "FP", "FN", "TN", "精确率", "召回率", "准确率", "F1"])
    for name, m in [
        ("v1 核心命令且非闲聊 · asr", v1_m),
        ("v2 任务导向主类 · asr", v2_m),
        ("v1 · pos cmd vs neg asr（理想）", v1_ideal),
        ("v2 · pos cmd vs neg asr（理想）", v2_ideal),
    ]:
        ws.append([name, m["TP"], m["FP"], m["FN"], m["TN"], m["精确率"], m["召回率"], m["准确率"], m["F1"]])
    ws.append([])
    ws.append(["说明", "v1 二分类 = 动作/状态/模式/复合 且 无闲聊标签（语气词会直接否决）。"])
    ws.append(["", "v2 任务导向 = 动作/状态/模式/复合/询问/生活陈述/省略指令。闲聊、播报、其他、空转写为非任务。"])
    style_header(ws)
    autosize(ws, [36, 8, 8, 8, 8, 12, 12, 12, 10])

    # 4 holdout
    ws = wb.create_sheet("灰区holdout")
    ws.append(["规则", "train contest", "test contest", "test RR", "test FRR", "test CER"])
    for h in hold_rows:
        ws.append([h["name"], h["train_C"], h["test_C"], h["test_RR"], h["test_FRR"], h["test_CER"]])
    ws.append([])
    ws.append(["τ_zh / τ_en（仅 train）", zt, et])
    ws.append(["灰区", "|score-τ|≤0.05"])
    ws.append(["口径", "真实 contest=0.5*RR+0.5*(1-含误拒CER)；分层 holdout 30%"])
    style_header(ws)
    autosize(ws, [28, 16, 14, 12, 12, 12])

    # 5 纠正示例
    ws = wb.create_sheet("v1误伤纠正示例")
    ws.append(["uid", "cmd_text", "asr_text", "v1标签", "v2主类"])
    for x in fixed_particle:
        ws.append([x["uid"], x["cmd"], x["asr"], x["v1"], x["v2"]])
    style_header(ws)
    autosize(ws, [12, 36, 40, 28, 16])

    ws = wb.create_sheet("复合命令回收示例")
    ws.append(["uid", "cmd_text", "v1标签", "v2主类"])
    for x in compound_gain:
        ws.append([x["uid"], x["cmd"], x["v1"], x["v2"]])
    style_header(ws)
    autosize(ws, [12, 55, 32, 14])

    ws = wb.create_sheet("短命令型负样本")
    ws.append(["uid", "asr_text", "v2主类", "长度"])
    for x in still_hard_neg:
        ws.append([x["uid"], x["asr"], x["v2"], x["len"]])
    ws.append([])
    ws.append(["说明", "这些 neg 在文本上就是合法短命令，范式无法拒；必须靠声纹。"])
    style_header(ws)
    autosize(ws, [14, 50, 16, 8])

    # 正负主类各 5 例
    ws = wb.create_sheet("v2正样本主类示例")
    ws.append(["主类", "uid", "cmd_text", "asr_text"])
    seen: dict[str, int] = Counter()
    for r in pos:
        p = r["v2_hyp"]["primary"]
        if seen[p] >= 4:
            continue
        seen[p] += 1
        ws.append([p, r["uid"], r.get("cmd_text"), r.get("asr_text")])
    style_header(ws)
    autosize(ws, [16, 12, 40, 44])

    ws = wb.create_sheet("v2负样本主类示例")
    ws.append(["主类", "uid", "wake_text", "asr_text"])
    seen = Counter()
    for r in neg:
        p = r["v2_hyp"]["primary"]
        if seen[p] >= 4:
            continue
        seen[p] += 1
        ws.append([p, r["uid"], r.get("wake_text"), r.get("asr_text")])
    style_header(ws)
    autosize(ws, [16, 12, 16, 50])

    ws = wb.create_sheet("结论")
    ws.append(["编号", "说明"])
    conclusions = [
        ["1", "可以完善，而且应该先改分类几何，而不是加词。v1 的 7 类是有用的句式清单，问题是多标签+语气词当类。"],
        ["2", f"v2 互斥后 pos 命令侧以动作为主，复合命令从 {v1_cmd.get('复合命令', 0)} 提到 {v2_cmd.get('复合命令', 0)}；闲聊不再吃掉「打开…吧」。"],
        ["3", f"纯文本分正负：v1 asr F1={v1_m['F1']} 召回 {v1_m['召回率']}；v2 F1={v2_m['F1']} 召回 {v2_m['召回率']}。精确率仍受「neg 里也有真命令」上限约束。"],
        ["4", "holdout 上「灰区+v1非命令」容易伤 FRR；「灰区+长度≥16」更稳。v2 非任务适合与长度合取（长且非任务），不要单独替代余弦。"],
        ["5", "带到赛方同场景集：迁移句式框架（动词、调到、疑问词、第一人称、话语标记），不要迁移设备名/品牌模式名。"],
        ["6", "下一步若还要涨分：灰区逻辑回归用 [score, len, task_oriented, n_verb]，在 train 折拟合，禁止全量重扫。"],
    ]
    for row in conclusions:
        ws.append(row)
    style_header(ws)
    autosize(ws, [8, 100])

    out_xlsx = root / "linguistic_pattern_analysis_v2.xlsx"
    wb.save(out_xlsx)
    print(json.dumps({
        "xlsx": str(out_xlsx),
        "v1_asr": v1_m,
        "v2_asr": v2_m,
        "overlap": overlap_v1_cmd,
        "particle_false_chat": particle_override,
        "agree": agree,
        "v2_cmd": dict(v2_cmd),
        "holdout": hold_rows,
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
